import os
import data_utils as utils
import numpy as np
import openpyxl
import pandas as pd

from sklearn.ensemble import IsolationForest
from sklearn.model_selection import train_test_split


""" 
########## Setting Variable Paths ########## 
"""
# Variable paths
dataset_base_path = '/home/li.zhiyuan/Desktop/Probabilistic Inference/Project/dataset/'
wesad_dataset = os.path.join(dataset_base_path, 'wesad', 'WESAD')


"""" 
########## Reading pickle files for all subjects ##########
"""
all_subj_pkl = utils.read_all_subj_pkl(wesad_dataset)  # a list of 15 dictionaries (one for each subject)


"""
########## Loop through multiple train_test_split ##########
"""
# declaring some global variables
idx = 1
file_name = ""
iso_forest_path = ""
b_avg_acc = 0
s_avg_acc = 0
percent_20 = 11000
percent_15 = 8400
percent_10 = 5600
percent_5 = 2800
percent_2_5 = 1400

test_range = 1000
n_est = 150
contam = 0.2
max_f = 32
rand_state = 42
fit_range = 1
train_test_split_range = 10

# create workbook object
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "iso_forest"

# add titles in the first row of each column
sheet.cell(row=1, column=1).value = "split_number"
sheet.cell(row=1, column=2).value = "fit_number"
sheet.cell(row=1, column=3).value = "n_estimators"
sheet.cell(row=1, column=4).value = "contamination"
sheet.cell(row=1, column=5).value = "max_features"
sheet.cell(row=1, column=6).value = "random_state"
sheet.cell(row=1, column=7).value = "baseline_accuracy"
sheet.cell(row=1, column=8).value = "stress_accuracy"

for t in range(train_test_split_range):
    print("This is split number {}".format(t))
    """
    ########## Splitting the dataset into train and test at 80:20 ratio ##########
    """
    train, test = train_test_split(all_subj_pkl, train_size=0.8)
    print("train, test : {}, {}".format(len(train), len(test)))


    """
    ########## Getting individual classes: baseline, stress, amusement, meditation ##########
    """
    baseline_data_train = utils.get_class_from_data(train, 'baseline')
    stress_data_train = utils.get_class_from_data(train, 'stress')
    amusement_data_train = utils.get_class_from_data(train, 'amusement')
    meditation_data_train = utils.get_class_from_data(train, 'meditation')

    # used for testing
    baseline_data_test = utils.get_class_from_data(test, 'baseline')
    stress_data_test = utils.get_class_from_data(test, 'stress')


    """
    ########## Cleaning the data - removing outliers ##########
    """
    df_baseline_train = pd.DataFrame(data=baseline_data_train,
                               columns=["c_ax", "c_ay", "c_az", "c_ecg", "c_emg", "c_eda", "c_temp", "c_resp"])
    print("Baseline train df shape = {}".format(df_baseline_train.shape))

    df_stress_train = pd.DataFrame(data=stress_data_train,
                                     columns=["c_ax", "c_ay", "c_az", "c_ecg", "c_emg", "c_eda", "c_temp", "c_resp"])
    print("Stress train df shape = {}".format(df_stress_train.shape))

    df_baseline_test = pd.DataFrame(data=baseline_data_test,
                                     columns=["c_ax", "c_ay", "c_az", "c_ecg", "c_emg", "c_eda", "c_temp", "c_resp"])
    print("Baseline test df shape = {}".format(df_baseline_test.shape))

    df_stress_test = pd.DataFrame(data=stress_data_test,
                                   columns=["c_ax", "c_ay", "c_az", "c_ecg", "c_emg", "c_eda", "c_temp", "c_resp"])
    print("Stress test df shape = {}".format(df_stress_test.shape))

    # Calculate the baseline IQR
    baseline_train_q1 = df_baseline_train.quantile(0.25)
    baseline_train_q3 = df_baseline_train.quantile(0.75)
    baseline_train_iqr = baseline_train_q3 - baseline_train_q1
    print("Baseline train IQR is {}".format(baseline_train_iqr))

    stress_train_q1 = df_stress_train.quantile(0.25)
    stress_train_q3 = df_stress_train.quantile(0.75)
    stress_train_iqr = stress_train_q3 - stress_train_q1
    print("Stress train IQR is {}".format(stress_train_iqr))

    baseline_test_q1 = df_baseline_test.quantile(0.25)
    baseline_test_q3 = df_baseline_test.quantile(0.75)
    baseline_test_iqr = baseline_test_q3 - baseline_test_q1
    print("Baseline test IQR is {}".format(baseline_test_iqr))

    stress_test_q1 = df_stress_test.quantile(0.25)
    stress_test_q3 = df_stress_test.quantile(0.75)
    stress_test_iqr = stress_test_q3 - stress_test_q1
    print("Stress test IQR is {}".format(stress_test_iqr))

    # Removing baseline outliers
    df_baseline_train_clean = df_baseline_train[
        ~((df_baseline_train < (baseline_train_q1 - 1.5 * baseline_train_iqr)) | (df_baseline_train > (baseline_train_q3 + 1.5 * baseline_train_iqr))).any(
            axis=1)]
    print("Baseline df train cleaned shape = {}".format(df_baseline_train_clean.shape))

    df_stress_train_clean = df_stress_train[
        ~((df_stress_train < (stress_train_q1 - 1.5 * stress_train_iqr)) | (
                    df_stress_train > (stress_train_q3 + 1.5 * stress_train_iqr))).any(
            axis=1)]
    print("Stress df train cleaned shape = {}".format(df_stress_train_clean.shape))

    df_baseline_test_clean = df_baseline_test[
        ~((df_baseline_test < (baseline_test_q1 - 1.5 * baseline_test_iqr)) | (
                    df_baseline_test > (baseline_test_q3 + 1.5 * baseline_test_iqr))).any(
            axis=1)]
    print("Baseline df test cleaned shape = {}".format(df_baseline_test_clean.shape))

    df_stress_test_clean = df_stress_test[
        ~((df_stress_test < (stress_test_q1 - 1.5 * stress_test_iqr)) | (
                df_stress_test > (stress_test_q3 + 1.5 * stress_test_iqr))).any(
            axis=1)]
    print("Stress df test cleaned shape = {}".format(df_stress_test_clean.shape))

    # Convert the dataframe to list
    print("Converting df_baseline_train_clean to list")
    baseline_data_train_clean = utils.df_to_list(df_baseline_train_clean)
    print("Converting df_stress_train_clean to list")
    stress_data_train_clean = utils.df_to_list(df_stress_train_clean)
    print("Converting df_baseline_test_clean to list")
    baseline_data_test_clean = utils.df_to_list(df_baseline_test_clean)
    print("Converting df_stress_test_clean to list")
    stress_data_test_clean = utils.df_to_list(df_stress_test_clean)

    """
    ########## Sliding window of 0.25 seconds ##########
    """
    # Readings taken at 700Hz --> 175 samples (at 0.25s window)

    # windows_b for baseline data train
    windows_b = utils.extract_windows(baseline_data_train_clean, 0.25)
    print("baseline data windows : {}".format(len(windows_b)))

    # windows_s for stress data train
    windows_s = utils.extract_windows(stress_data_train_clean, 0.25)
    print("stress data windows : {}".format(len(windows_s)))

    # windows_b_t for baseline data test
    windows_b_t = utils.extract_windows(baseline_data_test_clean, 0.25)
    print("baseline test data windows : {}".format(len(windows_b_t)))

    # windows_s_t for stress data test
    windows_s_t = utils.extract_windows(stress_data_test_clean, 0.25)
    print("stress test data windows : {}".format(len(windows_s_t)))


    """
    ########## Preparing training data ##########
    """
    temp_train_all = []
    for each in windows_b:
        temp_train_all.append(each)
    print("baseline train length : {}".format(len(temp_train_all)))

    for s in range(percent_20):
        temp_train_all.append(windows_s[s])
    print("stress + baseline train length : {}".format(len(temp_train_all)))


    """
    ########## Preparing testing data ##########
    """
    temp_test_all = []
    for i in range(test_range):
        temp_test_all.append(windows_b_t[i])
    print("baseline test length : {}".format(len(temp_test_all)))

    for j in range(test_range):
        temp_test_all.append(windows_s_t[j])
    print("stress + baseline test length : {}".format(len(temp_test_all)))


    """
    ########## Perform feature engineering on each of the windows ##########
    """
    # Extracting features for training data
    all_features_train = []
    for each in temp_train_all:
        features = utils.feature_extraction(each)
        all_features_train.append(features)
    print("temp_train length : {} -- with shape {}".format(len(all_features_train), np.array(all_features_train).shape))

    # Extracting features for testing data
    all_features_test = []
    for each in temp_test_all:
        features = utils.feature_extraction(each)
        all_features_test.append(features)
    print("temp_test length : {} -- with shape {}".format(len(all_features_test), np.array(all_features_test).shape))


    """
    ########## Fitting the model ##########
    """
    output_path = "/home/li.zhiyuan/Desktop/Probabilistic Inference/Project/output/"
    iso_forest_path = os.path.join(output_path, "isolation forest")
    if not os.path.exists(iso_forest_path):
        os.makedirs(iso_forest_path)

    file_name = "iso_forest_s_{}_f_{}_c_{}_r_{}_{}.xlsx".format(train_test_split_range, fit_range, contam, rand_state, idx)

    for f in range(fit_range):
        print("iteration number {}".format(f))
        isoForest = IsolationForest(n_estimators=n_est, contamination=contam, max_features=max_f, random_state=rand_state)
        isoForest.fit(all_features_train)

        print("model fitted")

        # fit_pred = isoForest.fit_predict(all_features_train)
        # print(fit_pred)

        pred = isoForest.predict(all_features_test)
        print(pred)

        count = 0
        for i in range(test_range):
            if pred[i] == 1:
                count += 1
        b_acc = count / test_range * 100
        b_avg_acc += b_acc
        print("baseline count is {} and acc is {}%".format(count, b_acc))

        s_count = 0
        for j in range(test_range + 1, test_range * 2):
            if pred[j] == -1:
                s_count += 1
        s_acc = s_count / test_range * 100
        s_avg_acc += s_acc
        print("stress count is {} and acc is {}%".format(s_count, s_acc))


        """
        ########## Save accuracy for each individual run ##########
        """
        sheet.cell(row=t*fit_range+f+2, column=1).value = t
        sheet.cell(row=t*fit_range+f+2, column=2).value = f
        sheet.cell(row=t*fit_range+f+2, column=3).value = n_est
        sheet.cell(row=t*fit_range+f+2, column=4).value = contam
        sheet.cell(row=t*fit_range+f+2, column=5).value = max_f
        sheet.cell(row=t*fit_range+f+2, column=6).value = rand_state
        sheet.cell(row=t*fit_range+f+2, column=7).value = b_acc
        sheet.cell(row=t*fit_range+f+2, column=8).value = s_acc

        file_path = os.path.join(iso_forest_path, file_name)
        wb.save(file_path)
        print("Saved split {} fit {} to excel file {}".format(t, f, file_path))


"""
########## Save the average accuracy to Excel ##########
"""
file_name_avg = "iso_forest_avg_s_{}_f_{}_c_{}_r_{}_{}.xlsx".format(train_test_split_range, fit_range, contam, rand_state, idx)

# create workbook object
wba = openpyxl.Workbook()
sheet = wba.active
sheet.title = "iso_forest_avg"

# add titles in the first row of each column
sheet.cell(row=1, column=1).value = "split_range"
sheet.cell(row=1, column=2).value = "fit_range"
sheet.cell(row=1, column=3).value = "n_estimators"
sheet.cell(row=1, column=4).value = "contamination"
sheet.cell(row=1, column=5).value = "max_features"
sheet.cell(row=1, column=6).value = "random_state"
sheet.cell(row=1, column=7).value = "avg_baseline_accuracy"
sheet.cell(row=1, column=8).value = "avg_stress_accuracy"

b_avg_acc = b_avg_acc / (train_test_split_range * fit_range)
print("average baseline acc is {}%".format(b_avg_acc))
s_avg_acc = s_avg_acc / (train_test_split_range * fit_range)
print("average stress acc is {}%".format(s_avg_acc))

sheet.cell(row=2, column=1).value = train_test_split_range
sheet.cell(row=2, column=2).value = fit_range
sheet.cell(row=2, column=3).value = n_est
sheet.cell(row=2, column=4).value = contam
sheet.cell(row=2, column=5).value = max_f
sheet.cell(row=2, column=6).value = rand_state
sheet.cell(row=2, column=7).value = b_avg_acc
sheet.cell(row=2, column=8).value = s_avg_acc

file_path_avg = os.path.join(iso_forest_path, file_name_avg)
wba.save(file_path_avg)
print("Saved average accuracy to excel file {}".format(file_path_avg))